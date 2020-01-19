import os
from xml.dom.minidom import parse

from openpyxl import load_workbook


def get_filenames(file_dir):
    '''
    获取所有xml的文件名
    :param file_dir:xml存放的文件路径
    :return:xml文件名列表
    '''
    for root, dirs, files in os.walk(file_dir):
        # print('files:', files)  # 当前路径下所有非目录子文件
        return files


def get_students_data(path, filenames):
    '''
    获取一个xml文件中的学生信息
    :param path:
    :param filenames:
    :return:
    '''
    stu_ls = []
    domTree = parse(f"{path}{filenames}")
    # 文档根元素
    rootNode = domTree.documentElement
    Students = rootNode.getElementsByTagName('Student')
    # 考试地点
    ExamRoomName = rootNode.getElementsByTagName('ExamRoomName')
    ExamRoom = ExamRoomName[0].childNodes[0].data
    # 考试时间日期
    ExamTime = rootNode.getElementsByTagName('ExamTime')
    exam_date_and_time = ExamTime[0].childNodes[0].data
    # 考试日期
    exam_date = exam_date_and_time.split(' ', 1)[0]
    # 考试时间
    exam_time = exam_date_and_time.split(' ', 1)[1]
    for Student in Students:
        ls = []
        # 学生的索引号
        # print(Student.getAttribute('index'))
        # 座位号
        Seat = Student.getElementsByTagName('Seat')[0]
        # print(Seat.nodeName, ":", Seat.childNodes[0].data)
        # 学号
        StudentCode = Student.getElementsByTagName('StudentCode')[0]
        # print(StudentCode.nodeName, ":", StudentCode.childNodes[0].data)
        # 姓名
        # RealName = Student.getElementsByTagName('RealName')[0]
        # print(RealName.nodeName, ":", RealName.childNodes[0].data)
        # 试卷号
        SubjectCode = Student.getElementsByTagName('SubjectCode')[0]
        # print(SubjectCode.nodeName, ":", SubjectCode.childNodes[0].data)
        # SubjectName = Student.getElementsByTagName('SubjectName')[0]
        # print(SubjectName.nodeName, ":", SubjectName.childNodes[0].data)
        # print(exam_date)
        # print(exam_time)
        ls.append(StudentCode.childNodes[0].data)
        ls.append(ExamRoom)
        ls.append(Seat.childNodes[0].data)
        ls.append(SubjectCode.childNodes[0].data)
        ls.append(exam_date)
        ls.append(exam_time)
        stu_ls.append(ls)
    return stu_ls


def get_students_exam_info_data(filenames, path):
    '''
    获取学生考试信息列表
    :return:
    '''
    stu_ls_ = []
    for filename in filenames:
        for i in get_students_data(path, filename):
            stu_ls_.append(i)
    return stu_ls_


def from_computer_course_info_get_students_data(path='计算机应用基础模板.xlsx', sheet_name='计算机应用基础模板'):
    '''
    获取“计算机应用基础”考生信息
    :param path: 计算机应用基础excel表所在路径
    :param sheet_name: 表名
    :return: 学生信息列表
    '''
    ls = []
    wb = load_workbook(path)
    rows = wb[sheet_name].max_row
    for i in range(2, rows + 1):
        ls_ = []
        # 座位号
        exam_seat = wb[sheet_name][f'A{i}'].value
        # 机房
        exam_room = wb[sheet_name][f'C{i}'].value
        # 考试日期时间
        exam_date_and_time = wb[sheet_name][f'D{i}'].value
        # 考试日期
        exam_date = exam_date_and_time.split(' ')[0]
        # 考试时间
        exam_time = exam_date_and_time.split(' ')[1]
        # 学号
        exam_code = wb[sheet_name][f'E{i}'].value
        # 姓名
        exam_name = wb[sheet_name][f'F{i}'].value
        ls_.append(exam_code)
        ls_.append(exam_room)
        ls_.append(exam_seat)
        ls_.append('1200')
        ls_.append(exam_date)
        ls_.append(exam_time)
        ls.append(ls_)
    return ls
