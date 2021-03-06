from lib.func import get_filenames, get_students_exam_info_data, \
    from_computer_course_info_get_students_data
from openpyxl import load_workbook
import PySimpleGUI as sg

layout = [[sg.Text('选择xml文件夹')], [sg.Input(), sg.FolderBrowse('浏览')],
          [sg.Text('选择"计算机应用基础"excel文件')], [sg.Input(), sg.FileBrowse('浏览')],
          [sg.Text('选择"考试通知单"excel文件')], [sg.Input(), sg.FileBrowse('浏览')],
          [sg.Text('处理进度')], [sg.ProgressBar(1000, orientation='h', size=(40, 20), key='progressbar')],
          [sg.Button('开始处理', size=(10, 1)), sg.Button('关闭', size=(10, 1))]]

window = sg.Window('合并考试通知单（openpyxl 版）', layout, icon='icon/puple128.ico')
# 进度条
progress_bar = window['progressbar']
while True:
    event, values = window.read()
    if event in (None, '关闭'):
        break
    try:
        # 获取指定文件夹下所有xml文件名列表
        path_xml = values[0] + '/'
        filenames = get_filenames(path_xml)[1:]
        zyk_datas = get_students_exam_info_data(filenames, path_xml)
    except Exception as e:
        print(e)
        sg.Popup('请检查选择的文件路径是否正确！', no_titlebar=True, keep_on_top=True, background_color='grey')
    try:
        # 获取计算机应用基础考生信息
        path_jsj = values[1]
        sheet_name = '计算机应用基础模板'
        jsj_datas = from_computer_course_info_get_students_data(path=path_jsj)
        print(f'计算机应用基础人数{len(jsj_datas)}')
        print(f'网考人数{len(zyk_datas)}')
        # 数据合并
        for data in jsj_datas:
            zyk_datas.append(data)
        print(f'合并后总人数{len(zyk_datas)}')
    except Exception as e:
        print(e)
        sg.Popup('请检查选择的文件格式是否正确，表名是否为"计算机应用基础模板"', no_titlebar=True, keep_on_top=True, background_color='grey')
    try:
        # 写入考试通知单
        path_tzd = values[2]
        wb = load_workbook(path_tzd)

        # 获取行数
        rows = wb['sheet1'].max_row
        for a_i in range(1, rows + 1):
            # 进度条
            # sg.OneLineProgressMeter('My Meter', a_i + 1, rows + 1, 'key', 'Optional message', orientation='h')
            progress_bar.UpdateBar(a_i + 1, max=rows + 1)
            print(f'-------程序进行到第{a_i}行-------')
            a_val = wb['sheet1'][f'A{a_i}'].value
            try:
                stu_id = a_val[24:37]  # 学号
                # 如果学号存在，从数据中遍历学生信息
                for i in range(len(zyk_datas) - 1, -1, -1):  # 倒叙遍历，为了可以删除找到的元素，缩短查找时间
                    stu_info = zyk_datas[i]
                    # 如果学号在信息表中找到
                    if stu_id == stu_info[0]:
                        # 遍历该学生所有考试科目，考试科目数量不会超过30
                        for j in range(2, 30):
                            # 获取通知单试卷号
                            sjh = wb['sheet1'][f'A{a_i + j}'].value
                            # 先判断sjh是否为'考点名称：秦皇岛电大'，
                            if sjh == '考点名称：秦皇岛电大':
                                # print(sjh)
                                break
                            # 如果试卷号在列表中
                            elif str(sjh) in stu_info:
                                # 写入考场号
                                wb['sheet1'][f'D{a_i + j}'].value = stu_info[1]
                                # 写入座位号
                                wb['sheet1'][f'E{a_i + j}'].value = stu_info[2]
                                # 写入考试日期
                                wb['sheet1'][f'F{a_i + j}'].value = stu_info[4]
                                # 写入考试时间
                                wb['sheet1'][f'G{a_i + j}'].value = stu_info[5]
                                # print(stu_id)
                                print(f'{stu_id}试卷号{sjh}写入完毕！')
                                # 移除找到的元素，缩小查找范围
                                zyk_datas.pop(i)
                    # print(len(zyk_datas))
            except Exception as e:
                print(e)
                continue
        wb.save(path_tzd)
        wb.close()
        sg.Popup('程序执行完毕！', no_titlebar=True, keep_on_top=True, background_color='grey')
    except Exception as e:
        print(e)
        sg.Popup('请检查选"考试通知单"所在路径或文件是否正确！', no_titlebar=True, keep_on_top=True, background_color='grey')

window.close()
