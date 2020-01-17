from merge_exam_info.lib.func import get_filenames, get_students_exam_info_data
from openpyxl import load_workbook

# 中央开学生考试数据
# path = '签到表/中央开网考/data/'
path = 'xml/'
# 获取指定文件夹下所有xml文件名列表
filenames = get_filenames(path)[1:]
zyk_datas = get_students_exam_info_data(filenames, path)

for i in zyk_datas:
    print(i)

print(len(zyk_datas))

wb = load_workbook('tzd.xlsx')

# 获取行数
rows = wb['sheet1'].max_row
for a_i in range(1, rows + 1):
    print(f'-------程序进行到第{a_i}行-------')
    a_val = wb['sheet1'][f'A{a_i}'].value
    try:
        stu_id = a_val[24:37]  # 学号
        # 如果学号存在，从数据中遍历学生信息
        for i in range(len(zyk_datas) - 1, -1, -1):  # 倒叙遍历，为了可以删除找到的元素，缩短查找时间
            stu_info = zyk_datas[i]
            # 如果学号在信息表中找到
            if stu_id == stu_info[0]:
                # 遍历该学生所有考试科目，考试科目数量不会超过30
                for j in range(2, 30):
                    # 获取通知单试卷号
                    sjh = wb['sheet1'][f'A{a_i + j}'].value
                    # 先判断sjh是否为'考点名称：秦皇岛电大'，
                    if sjh == '考点名称：秦皇岛电大':
                        # print(sjh)
                        break
                    # 如果试卷号在列表中
                    elif str(sjh) in stu_info:
                        # 写入考场号
                        wb['sheet1'][f'D{a_i + j}'].value = stu_info[1]
                        # 写入座位号
                        wb['sheet1'][f'E{a_i + j}'].value = stu_info[2]
                        # 写入考试日期
                        wb['sheet1'][f'F{a_i + j}'].value = stu_info[4]
                        # 写入考试时间
                        wb['sheet1'][f'G{a_i + j}'].value = stu_info[5]
                        # print(stu_id)
                        print(f'{stu_id}试卷号{sjh}写入完毕！')
                        # 移除找到的元素，缩小查找范围
                        zyk_datas.pop(i)
            # print(len(zyk_datas))
    except Exception as e:
        print(e)
        continue
wb.save('tzd.xlsx')
wb.close()
